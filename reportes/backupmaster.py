import os
import pandas as pd
from config.db import SessionLocal
from sqlalchemy import text
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
from datetime import datetime

logger = logging.getLogger("uvicorn")
load_dotenv()

def calcular_tiempo_diferencia(fecha_inicio, fecha_fin):
    """
    Calcula la diferencia entre dos fechas y retorna en formato HH:MM:SS
    Args:
        fecha_inicio: datetime de inicio
        fecha_fin: datetime de fin
    Returns:
        String en formato HH:MM:SS
    """
    if not fecha_inicio or not fecha_fin:
        return "00:00:00"
    
    try:
        diferencia = fecha_fin - fecha_inicio
        segundos_totales = int(diferencia.total_seconds())
        
        if segundos_totales < 0:
            return "00:00:00"
        
        # Convertir a HH:MM:SS
        horas = segundos_totales // 3600
        minutos = (segundos_totales % 3600) // 60
        segundos = segundos_totales % 60
        
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
    except:
        return "00:00:00"

def obtener_todas_tareas_finalizadas():
    """
    Obtiene todas las tareas que tienen fecha_fin.
    Returns:
        Lista de diccionarios con los datos de las tareas
    """
    session = SessionLocal()
    try:
        query = text("""
            SELECT 
                t.id_tarea,
                t.fecha_inicio,
                t.fecha_fin,
                t.tiempo_total,
                t.numero_op,
                t.numero_plano,
                t.id_sector,
                t.id_producto,
                t.nombre_labor,
                t.id_operario_seleccionado,
                t.id_usuario_logeado,
                sec.nombre AS nombre_sector,
                prod.nombre AS nombre_producto,
                t.nombre_operario_seleccionado,
                t.apellido_operario_seleccionado,
                t.apellido_usuario_logeado,
                t.nombre_usuario_logeado
            FROM tareas t
            LEFT JOIN sectores sec ON t.id_sector = sec.id_sector
            LEFT JOIN productos prod ON t.id_producto = prod.id_producto
            WHERE t.fecha_fin IS NOT NULL
            ORDER BY t.fecha_inicio
        """)
        result = session.execute(query)
        rows = result.fetchall()
        
        if not rows:
            return []
        
        tareas = []
        for row in rows:
            datos = {
                "id": row[0],
                "fecha_inicio": row[1],
                "fecha_fin": row[2],
                "tiempo_total": row[3],
                "numero_op": row[4],
                "numero_plano": row[5],
                "id_sector": row[6],
                "id_producto": row[7],
                "nombre_labor": row[8],
                "id_operario": row[9],
                "id_usuario_creador": row[10],
                "nombre_sector": row[11],
                "nombre_producto": row[12],
                "nombre_operario": row[13],
                "apellido_operario": row[14],
                "apellido_usuario": row[15],
                "nombre_usuario": row[16]
            }
            tareas.append(datos)
        return tareas
    finally:
        session.close()

def export_todas_tareas_finalizadas_to_excel(file_path):
    """
    Genera un Excel con el resumen de todas las tareas finalizadas.
    Args:
        file_path: Ruta del archivo Excel a generar
    Returns:
        True si se generó correctamente, False si no hay datos
    """
    logger.info("Exportando datos de todas las tareas finalizadas")
    
    tareas = obtener_todas_tareas_finalizadas()
    
    if not tareas:
        logger.info("No se encontraron tareas finalizadas")
        return False
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte | Master"
        
        logo_path = os.path.join(os.path.dirname(__file__), "..", "imagenes", "cremonarecort.png")
        
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 31.5
            img.width = 126
            ws.add_image(img, "E3")
        
        ws.merge_cells("A1:K1")
        ws["A1"] = "REPORTE MASTER DE TAREAS"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws["A3"] = ""
        ws["A4"] = ""

        # Espacios para el logo

        ws.append([])
        
        headers = [
            "Fecha de inicio de la tarea\n[YYYY-MM-DD HH:MM:SS]",
            "Fecha de finalización de la tarea\n[YYYY-MM-DD HH:MM:SS]",
            "Tiempo bruto [HH:MM:SS]",
            "Tiempo neto [HH:MM:SS]",
            "OP",
            "Plano",
            "Sector",
            "Producto",
            "Labor",
            "Operario",
            "Creador de la tarea"
        ]
        
        ws.append(headers)
        first_table_first_row = ws.max_row
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=first_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws.row_dimensions[first_table_first_row].height = 45
        
        for datos_tarea in tareas:
            tiempo_bruto = calcular_tiempo_diferencia(datos_tarea["fecha_inicio"], datos_tarea["fecha_fin"])
            
            fecha_inicio_str = datos_tarea["fecha_inicio"].strftime("%Y-%m-%d %H:%M:%S") if datos_tarea["fecha_inicio"] else ""
            fecha_fin_str = datos_tarea["fecha_fin"].strftime("%Y-%m-%d %H:%M:%S") if datos_tarea["fecha_fin"] else ""
            
            tiempo_neto = datos_tarea["tiempo_total"] or "00:00:00"
            
            operario_nombre = ""
            if datos_tarea["apellido_operario"] and datos_tarea["nombre_operario"]:
                operario_nombre = f"{datos_tarea['apellido_operario']} {datos_tarea['nombre_operario']}"
            
            usuario_nombre = ""
            if datos_tarea["apellido_usuario"] and datos_tarea["nombre_usuario"]:
                usuario_nombre = f"{datos_tarea['apellido_usuario']} {datos_tarea['nombre_usuario']}"
            
            ws.append([
                fecha_inicio_str,
                fecha_fin_str,
                tiempo_bruto,
                tiempo_neto,
                datos_tarea["numero_op"] or "",
                datos_tarea["numero_plano"] or "",
                datos_tarea["nombre_sector"] or "",
                datos_tarea["nombre_producto"] or "",
                datos_tarea["nombre_labor"] or "",
                operario_nombre,
                usuario_nombre
            ])
        
        first_table_last_row = ws.max_row
        first_table = Table(displayName="ReporteMasterTareas", ref=f"A{first_table_first_row}:K{first_table_last_row}")
        first_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        first_table.tableStyleInfo = first_style
        ws.add_table(first_table)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        
        wb.save(file_path)
        logger.info(f"Excel generado exitosamente en: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error al generar el Excel: {str(e)}")
        return False
