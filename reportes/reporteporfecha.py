import os
import pandas as pd
from config.db import SessionLocal
from sqlalchemy import text
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
from datetime import datetime, timedelta

logger = logging.getLogger("uvicorn")
load_dotenv()

def formatear_header_multilinea(header_text, max_chars_por_linea=20, max_lineas=2):
    """
    Divide un encabezado largo en múltiples líneas si es necesario.
    Args:
        header_text: El texto del encabezado
        max_chars_por_linea: Máximo de caracteres por línea
        max_lineas: Máximo número de líneas (máximo 1 salto = 2 líneas)
    Returns:
        String con saltos de línea donde sea apropiado
    """
    if len(header_text) <= max_chars_por_linea:
        return header_text
    
    palabras = header_text.split(' ')
    if len(palabras) == 1:
        if '[' in header_text and ']' in header_text:
            parte1 = header_text[:header_text.find('[')].strip()
            parte2 = header_text[header_text.find('['):].strip()
            return f"{parte1}\n{parte2}"
        else:
            return header_text
    
    lineas = []
    linea_actual = ""
    
    for palabra in palabras:
        if len(linea_actual + " " + palabra) <= max_chars_por_linea:
            if linea_actual:
                linea_actual += " " + palabra
            else:
                linea_actual = palabra
        else:
            if linea_actual:
                lineas.append(linea_actual)
                linea_actual = palabra
                if len(lineas) >= max_lineas - 1:
                    palabras_restantes = palabras[palabras.index(palabra):]
                    linea_actual = " ".join(palabras_restantes)
                    break
            else:
                lineas.append(palabra)
    
    if linea_actual:
        lineas.append(linea_actual)
    
    if len(lineas) > max_lineas:
        lineas = lineas[:max_lineas]
    
    return "\n".join(lineas)

def calcular_tiempo_diferencia(fecha_inicio, fecha_fin):
    """
    Calcula la diferencia entre dos fechas y retorna en formato HH:MM:SS
    Args:
        fecha_inicio: datetime de inicio
        fecha_fin: datetime de fin
    Returns:
        String en formato HH:MM:SS
    """
    if not fecha_inicio or not fecha_fin:
        return "00:00:00"
    
    try:
        diferencia = fecha_fin - fecha_inicio
        segundos_totales = int(diferencia.total_seconds())
        
        if segundos_totales < 0:
            return "00:00:00"
        
        # Convertir a HH:MM:SS
        horas = segundos_totales // 3600
        minutos = (segundos_totales % 3600) // 60
        segundos = segundos_totales % 60
        
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
    except:
        return "00:00:00"

def obtener_tareas_por_fecha(fecha_inicio, fecha_fin):
    """
    Obtiene todas las tareas en un rango de fechas específico.
    Args:
        fecha_inicio: Fecha de inicio (formato YYYY-MM-DD)
        fecha_fin: Fecha de fin (formato YYYY-MM-DD)
    Returns:
        Lista de diccionarios con los datos de las tareas
    """
    session = SessionLocal()
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        
        query = text("""
            SELECT 
                t.id_tarea,
                t.fecha_inicio,
                t.fecha_fin,
                t.tiempo_total,
                t.numero_op,
                t.numero_plano,
                t.id_sector,
                t.id_producto,
                t.nombre_labor,
                t.id_operario_seleccionado,
                t.id_usuario_logeado,
                sec.nombre AS nombre_sector,
                prod.nombre AS nombre_producto,
                t.nombre_operario_seleccionado,
                t.apellido_operario_seleccionado,
                t.apellido_usuario_logeado,
                t.nombre_usuario_logeado
            FROM tareas t
            LEFT JOIN sectores sec ON t.id_sector = sec.id_sector
            LEFT JOIN productos prod ON t.id_producto = prod.id_producto
            WHERE t.fecha_fin >= :fecha_inicio AND t.fecha_fin <= :fecha_fin
            ORDER BY t.fecha_fin
        """)
        result = session.execute(query, {"fecha_inicio": fecha_inicio_dt, "fecha_fin": fecha_fin_dt})
        rows = result.fetchall()
        
        if not rows:
            return []
        
        tareas = []
        for row in rows:
            datos = {
                "id": row[0],
                "fecha_inicio": row[1],
                "fecha_fin": row[2],
                "tiempo_total": row[3],
                "numero_op": row[4],
                "numero_plano": row[5],
                "id_sector": row[6],
                "id_producto": row[7],
                "nombre_labor": row[8],
                "id_operario": row[9],
                "id_usuario_creador": row[10],
                "nombre_sector": row[11],
                "nombre_producto": row[12],
                "nombre_operario": row[13],
                "apellido_operario": row[14],
                "apellido_usuario": row[15],
                "nombre_usuario": row[16]
            }
            tareas.append(datos)
        return tareas
    finally:
        session.close()

def export_tareas_por_fecha_to_excel(file_path, fecha_inicio, fecha_fin):
    """
    Genera un Excel con el resumen de todas las tareas en un rango de fechas.
    Args:
        file_path: Ruta del archivo Excel a generar
        fecha_inicio: Fecha de inicio (formato YYYY-MM-DD)
        fecha_fin: Fecha de fin (formato YYYY-MM-DD)
    Returns:
        True si se generó correctamente, False si no hay datos
    """
    logger.info(f"Exportando datos de tareas para rango de fechas: {fecha_inicio} a {fecha_fin}")
    
    tareas = obtener_tareas_por_fecha(fecha_inicio, fecha_fin)
    
    if not tareas:
        logger.info(f"No se encontraron datos para el rango de fechas: {fecha_inicio} a {fecha_fin}")
        return False
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte | Fecha"
        
        logo_path = os.path.join(os.path.dirname(__file__), "..", "imagenes", "cremonarecort.png")
        
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 31.5
            img.width = 126
            ws.add_image(img, "K3")
        
        ws.merge_cells("A1:K1")
        ws["A1"] = "REPORTE DE TAREAS POR FECHA"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws["A3"] = "Fecha de inicio filtrado:"
        ws["A3"].font = Font(size=12, bold=True)
        ws["B3"] = fecha_inicio
        
        ws["A4"] = "Fecha de fin filtrado:"
        ws["A4"].font = Font(size=12, bold=True)
        ws["B4"] = fecha_fin
        
        ws.append([])
        
        headers = [
            "Fecha de inicio de la tarea\n[YYYY-MM-DD HH:MM:SS]",
            "Fecha de finalización de la tarea\n[YYYY-MM-DD HH:MM:SS]",
            "Tiempo bruto [HH:MM:SS]",
            "Tiempo neto [HH:MM:SS]",
            "OP",
            "Plano",
            "Sector",
            "Producto",
            "Labor",
            "Operario",
            "Creador de la tarea"
        ]
        
        ws.append(headers)
        first_table_first_row = ws.max_row
        
        # Aplicar formato a los encabezados
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=first_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws.row_dimensions[first_table_first_row].height = 45
        
        for datos_tarea in tareas:
            tiempo_bruto = calcular_tiempo_diferencia(datos_tarea["fecha_inicio"], datos_tarea["fecha_fin"])
            
            fecha_inicio_str = datos_tarea["fecha_inicio"].strftime("%Y-%m-%d %H:%M:%S") if datos_tarea["fecha_inicio"] else ""
            fecha_fin_str = datos_tarea["fecha_fin"].strftime("%Y-%m-%d %H:%M:%S") if datos_tarea["fecha_fin"] else ""
            
            tiempo_neto = datos_tarea["tiempo_total"] or "00:00:00"
            
            operario_nombre = ""
            if datos_tarea["apellido_operario"] and datos_tarea["nombre_operario"]:
                operario_nombre = f"{datos_tarea['apellido_operario']} {datos_tarea['nombre_operario']}"
            
            usuario_nombre = ""
            if datos_tarea["apellido_usuario"] and datos_tarea["nombre_usuario"]:
                usuario_nombre = f"{datos_tarea['apellido_usuario']} {datos_tarea['nombre_usuario']}"
            
            ws.append([
                fecha_inicio_str,
                fecha_fin_str,
                tiempo_bruto,
                tiempo_neto,
                datos_tarea["numero_op"] or "",
                datos_tarea["numero_plano"] or "",
                datos_tarea["nombre_sector"] or "",
                datos_tarea["nombre_producto"] or "",
                datos_tarea["nombre_labor"] or "",
                operario_nombre,
                usuario_nombre
            ])
        
        first_table_last_row = ws.max_row
        first_table = Table(displayName="ReporteTareasPorFecha", ref=f"A{first_table_first_row}:K{first_table_last_row}")
        first_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        first_table.tableStyleInfo = first_style
        ws.add_table(first_table)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        
        # Guardar el workbook
        wb.save(file_path)
        logger.info(f"Excel generado exitosamente en: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error al generar el Excel: {str(e)}")
        return False
