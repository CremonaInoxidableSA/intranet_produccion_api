import os
import pandas as pd
from config.db import SessionLocal
from sqlalchemy import text
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
from datetime import datetime

logger = logging.getLogger("uvicorn")
load_dotenv()
TIEMPO_EXCEL = int(os.getenv("TIEMPO_EXCEL", 1))

def formatear_header_multilinea(header_text, max_chars_por_linea=20, max_lineas=2):
    """
    Divide un encabezado largo en múltiples líneas si es necesario.
    Args:
        header_text: El texto del encabezado
        max_chars_por_linea: Máximo de caracteres por línea
        max_lineas: Máximo número de líneas (máximo 1 salto = 2 líneas)
    Returns:
        String con saltos de línea donde sea apropiado
    """
    if len(header_text) <= max_chars_por_linea:
        return header_text
    
    # Buscar espacios o caracteres de separación para dividir
    palabras = header_text.split(' ')
    if len(palabras) == 1:
        # Si es una sola palabra muy larga, buscar otros separadores
        if '[' in header_text and ']' in header_text:
            # Dividir antes del corchete
            parte1 = header_text[:header_text.find('[')].strip()
            parte2 = header_text[header_text.find('['):].strip()
            return f"{parte1}\n{parte2}"
        else:
            return header_text  # No dividir si no hay separadores naturales
    
    # Dividir en líneas manteniendo palabras completas
    lineas = []
    linea_actual = ""
    
    for palabra in palabras:
        if len(linea_actual + " " + palabra) <= max_chars_por_linea:
            if linea_actual:
                linea_actual += " " + palabra
            else:
                linea_actual = palabra
        else:
            if linea_actual:
                lineas.append(linea_actual)
                linea_actual = palabra
                # Si ya tenemos el máximo de líneas, agregar el resto a la última línea
                if len(lineas) >= max_lineas - 1:
                    # Agregar todas las palabras restantes a la línea actual
                    palabras_restantes = palabras[palabras.index(palabra):]
                    linea_actual = " ".join(palabras_restantes)
                    break
            else:
                lineas.append(palabra)
    
    if linea_actual:
        lineas.append(linea_actual)
    
    # Asegurar que no excedamos el máximo de líneas
    if len(lineas) > max_lineas:
        lineas = lineas[:max_lineas]
    
    return "\n".join(lineas)

def normalizar_tiempo(tiempo_str):
    """Convierte 'mm:ss' o 'hh:mm:ss' a 'hh:mm:ss'."""
    if not isinstance(tiempo_str, str):
        tiempo_str = str(tiempo_str)
    partes = tiempo_str.split(":")
    if len(partes) == 2:
        return f"00:{partes[0].zfill(2)}:{partes[1].zfill(2)}"
    elif len(partes) == 3:
        return f"{partes[0].zfill(2)}:{partes[1].zfill(2)}:{partes[2].zfill(2)}"
    else:
        try:
            segundos = int(float(tiempo_str))
            return str(pd.to_timedelta(segundos, unit="s"))
        except Exception:
            return "00:00:00"

def tiempo_a_minutos(tiempo_str):
    """Convierte tiempo en formato 'mm:ss' o 'hh:mm:ss' a minutos."""
    if not tiempo_str or not isinstance(tiempo_str, str):
        return 0
    
    try:
        partes = tiempo_str.split(":")
        if len(partes) == 2:  # mm:ss
            return int(partes[0]) + (int(partes[1]) / 60)
        elif len(partes) == 3:  # hh:mm:ss
            return int(partes[0]) * 60 + int(partes[1]) + (int(partes[2]) / 60)
        else:
            return 0
    except:
        return 0

def tiempo_a_mm_ss(tiempo_str):
    """Convierte tiempo en formato 'mm:ss' o 'hh:mm:ss' a formato 'MM:SS'."""
    if not tiempo_str or not isinstance(tiempo_str, str):
        return "00:00"
    
    try:
        partes = tiempo_str.split(":")
        if len(partes) == 2:  # mm:ss
            return f"{int(partes[0]):02d}:{int(partes[1]):02d}"
        elif len(partes) == 3:  # hh:mm:ss
            # Convertir hh:mm:ss a total de minutos y segundos
            total_minutos = int(partes[0]) * 60 + int(partes[1])
            segundos = int(partes[2])
            return f"{total_minutos:02d}:{segundos:02d}"
        else:
            return "00:00"
    except:
        return "00:00"

def calcular_tiempo_util_mm_ss(tiempo_desmolde_str, tiempo_pausado_str):
    """Calcula el tiempo útil restando tiempo pausado del tiempo de desmolde, retorna en MM:SS."""
    if not tiempo_desmolde_str:
        return "00:00"
    
    tiempo_pausado_str = tiempo_pausado_str or "00:00"
    
    try:
        # Convertir tiempo de desmolde a segundos totales
        partes_desmolde = tiempo_desmolde_str.split(":")
        segundos_desmolde = 0
        if len(partes_desmolde) == 2:  # mm:ss
            segundos_desmolde = int(partes_desmolde[0]) * 60 + int(partes_desmolde[1])
        elif len(partes_desmolde) == 3:  # hh:mm:ss
            segundos_desmolde = int(partes_desmolde[0]) * 3600 + int(partes_desmolde[1]) * 60 + int(partes_desmolde[2])
        
        # Convertir tiempo pausado a segundos totales
        partes_pausado = tiempo_pausado_str.split(":")
        segundos_pausado = 0
        if len(partes_pausado) == 2:  # mm:ss
            segundos_pausado = int(partes_pausado[0]) * 60 + int(partes_pausado[1])
        elif len(partes_pausado) == 3:  # hh:mm:ss
            segundos_pausado = int(partes_pausado[0]) * 3600 + int(partes_pausado[1]) * 60 + int(partes_pausado[2])
        
        # Calcular tiempo útil
        segundos_util = max(0, segundos_desmolde - segundos_pausado)
        
        # Convertir a MM:SS
        minutos = segundos_util // 60
        segundos = segundos_util % 60
        
        return f"{minutos:02d}:{segundos:02d}"
        
    except:
        return "00:00"

def calcular_tiempo_entre_ciclos(fecha_inicio_actual, fecha_fin_anterior):
    """
    Calcula el tiempo entre el inicio del ciclo actual y la finalización del ciclo anterior.
    Args:
        fecha_inicio_actual: Fecha de inicio del ciclo actual (datetime)
        fecha_fin_anterior: Fecha de fin del ciclo anterior (datetime)
    Returns:
        String en formato HH:MM:SS
    """
    if not fecha_inicio_actual or not fecha_fin_anterior:
        return "00:00:00"
    
    try:
        # Calcular la diferencia en segundos
        diferencia = fecha_inicio_actual - fecha_fin_anterior
        segundos_totales = int(diferencia.total_seconds())
        
        # Si la diferencia es negativa o muy grande (más de 24 horas), retornar 00:00:00
        if segundos_totales < 0 or segundos_totales > 86400:  # 24 horas = 86400 segundos
            return "00:00:00"
        
        # Convertir a HH:MM:SS
        horas = segundos_totales // 3600
        minutos = (segundos_totales % 3600) // 60
        segundos = segundos_totales % 60
        
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
        
    except:
        return "00:00:00"

def obtener_id_recetario_por_fecha(fecha):
    session = SessionLocal()
    try:
        query = text("""
            SELECT DISTINCT rxc.id_recetario
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
        """)
        result = session.execute(query, {"fecha": fecha})
        ids = [row[0] for row in result.fetchall()]
        return ids
    finally:
        session.close()

def obtener_codigos_producto_por_ids_recetario(id_recetarios):
    if not id_recetarios:
        return {}
    
    session = SessionLocal()
    try:
        # Convertir la lista a una cadena separada por comas para la consulta SQL
        ids_str = ", ".join(str(id) for id in id_recetarios)
        
        query = text(f"""
            SELECT id, codigoProducto
            FROM recetario
            WHERE id IN ({ids_str})
        """)
        
        result = session.execute(query)
        # Crear un diccionario {id_recetario: codigoProducto}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_cantidad_niveles_receta_por_ids_recetario(id_recetarios):
    """
    Obtiene la cantidad de niveles de la receta para cada id_recetario.
    Args:
        id_recetarios: Lista de IDs de recetarios
    Returns:
        Diccionario {id_recetario: cantidadNiveles}
    """
    if not id_recetarios:
        return {}
    
    session = SessionLocal()
    try:
        # Convertir la lista a una cadena separada por comas para la consulta SQL
        ids_str = ", ".join(str(id) for id in id_recetarios)
        
        query = text(f"""
            SELECT id, cantidadNiveles
            FROM recetario
            WHERE id IN ({ids_str})
        """)
        
        result = session.execute(query)
        # Crear un diccionario {id_recetario: cantidadNiveles}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_cantidad_ciclos_por_recetario(fecha):
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, COUNT(DISTINCT rxc.id_ciclo_desmoldeo) as cantidad_ciclos
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            AND cd.estadoMaquina != 'CANCELADO AL INICIAR'
            GROUP BY rxc.id_recetario
        """)
        result = session.execute(query, {"fecha": fecha})
        # Crear un diccionario {id_recetario: cantidad_ciclos}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_ciclos_cancelados_por_recetario(fecha):
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, COUNT(DISTINCT rxc.id_ciclo_desmoldeo) as ciclos_cancelados
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            AND cd.estadoMaquina = 'CANCELADO'
            GROUP BY rxc.id_recetario
        """)
        result = session.execute(query, {"fecha": fecha})
        # Crear un diccionario {id_recetario: ciclos_cancelados}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_peso_total_por_recetario(fecha):
    """Obtiene el peso total desmoldado para cada id_recetario en una fecha específica."""
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, SUM(cd.pesoDesmoldado) as peso_total
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            GROUP BY rxc.id_recetario
        """)
        result = session.execute(query, {"fecha": fecha})
        # Crear un diccionario {id_recetario: peso_total}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_tiempo_total_por_recetario(fecha):
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, cd.tiempoDesmolde, cd.tiempoPausado
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
        """)
        result = session.execute(query, {"fecha": fecha})
        
        # Agrupar tiempos por id_recetario
        tiempos_por_recetario = {}
        for row in result.fetchall():
            id_recetario = row[0]
            tiempo_desmolde_str = row[1]
            tiempo_pausado_str = row[2] or "00:00"  # Si es NULL, usar "00:00"
            
            if id_recetario not in tiempos_por_recetario:
                tiempos_por_recetario[id_recetario] = []
            
            tiempos_por_recetario[id_recetario].append({
                'desmolde': tiempo_desmolde_str,
                'pausado': tiempo_pausado_str
            })
        
        # Calcular el tiempo total para cada recetario
        totales = {}
        for id_recetario, tiempos in tiempos_por_recetario.items():
            segundos_totales = 0
            for tiempo_info in tiempos:
                # Convertir tiempo de desmolde a segundos
                tiempo_desmolde = tiempo_info['desmolde']
                partes_desmolde = tiempo_desmolde.split(":")
                segundos_desmolde = 0
                if len(partes_desmolde) == 2:  # mm:ss
                    segundos_desmolde = int(partes_desmolde[0]) * 60 + int(partes_desmolde[1])
                elif len(partes_desmolde) == 3:  # hh:mm:ss
                    segundos_desmolde = int(partes_desmolde[0]) * 3600 + int(partes_desmolde[1]) * 60 + int(partes_desmolde[2])
                
                # Convertir tiempo pausado a segundos
                tiempo_pausado = tiempo_info['pausado']
                partes_pausado = tiempo_pausado.split(":")
                segundos_pausado = 0
                if len(partes_pausado) == 2:  # mm:ss
                    segundos_pausado = int(partes_pausado[0]) * 60 + int(partes_pausado[1])
                elif len(partes_pausado) == 3:  # hh:mm:ss
                    segundos_pausado = int(partes_pausado[0]) * 3600 + int(partes_pausado[1]) * 60 + int(partes_pausado[2])
                
                # Restar tiempo pausado del tiempo de desmolde y sumar al total
                tiempo_efectivo = segundos_desmolde - segundos_pausado
                # Asegurar que no sea negativo
                tiempo_efectivo = max(0, tiempo_efectivo)
                segundos_totales += tiempo_efectivo
            
            # Convertir segundos totales a formato hh:mm:ss
            horas = segundos_totales // 3600
            minutos = (segundos_totales % 3600) // 60
            segundos = segundos_totales % 60
            
            totales[id_recetario] = f"{horas:02d}:{minutos:02d}:{segundos:02d}"
        
        return totales
    finally:
        session.close()

def obtener_niveles_desmoldados_por_recetario(fecha):
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, SUM(rxc.cantidadNivelesFinalizado) as niveles_desmoldados
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            GROUP BY rxc.id_recetario
        """)
        result = session.execute(query, {"fecha": fecha})
        # Crear un diccionario {id_recetario: niveles_desmoldados}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_segundos_por_nivel_por_recetario(tiempos_totales, niveles_desmoldados):
    """
    Calcula los segundos por nivel para cada recetario.
    Args:
        tiempos_totales: Diccionario {id_recetario: "hh:mm:ss"}
        niveles_desmoldados: Diccionario {id_recetario: cantidad_niveles}
    Returns:
        Diccionario {id_recetario: "X seg"}
    """
    segundos_por_nivel = {}
    
    for id_recetario, tiempo_total in tiempos_totales.items():
        niveles = niveles_desmoldados.get(id_recetario, 0)
        
        try:
            # Convertir tiempo_total (hh:mm:ss) a segundos totales
            tiempo_partes = tiempo_total.split(":")
            segundos_totales = int(tiempo_partes[0]) * 3600 + int(tiempo_partes[1]) * 60 + int(tiempo_partes[2])
            
            # Calcular segundos por nivel (si hay niveles)
            if niveles > 0:
                segundos_por_nivel[id_recetario] = int(segundos_totales / niveles)
            else:
                segundos_por_nivel[id_recetario] = "0"
        except Exception:
            segundos_por_nivel[id_recetario] = "0"
    
    return segundos_por_nivel

def calcular_porcentaje_fallas(ciclos_totales, ciclos_cancelados):
    """
    Calcula el porcentaje de fallas para cada recetario.
    Args:
        ciclos_totales: Diccionario {id_recetario: total_ciclos}
        ciclos_cancelados: Diccionario {id_recetario: ciclos_cancelados}
    Returns:
        Diccionario {id_recetario: porcentaje_fallas}
    """
    porcentajes = {}
    
    for id_recetario, total in ciclos_totales.items():
        cancelados = ciclos_cancelados.get(id_recetario, 0)
        
        if total > 0:
            porcentajes[id_recetario] = round((cancelados / total) * 100, 2)
        else:
            porcentajes[id_recetario] = 0
    
    return porcentajes

def obtener_niveles_fallados_por_recetario(fecha):
    """
    Obtiene los niveles fallados (cuenta de ciclos cancelados) para cada recetario.
    Args:
        fecha: Fecha en formato "YYYY-MM-DD"
    Returns:
        Diccionario {id_recetario: niveles_fallados}
    """
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, COUNT(DISTINCT rxc.id_ciclo_desmoldeo) as niveles_fallados
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            AND cd.estadoMaquina = 'CANCELADO'
            GROUP BY rxc.id_recetario
        """)
        result = session.execute(query, {"fecha": fecha})
        # Crear un diccionario {id_recetario: niveles_fallados}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def obtener_niveles_ciclados_neto_por_recetario(fecha):
    """
    Obtiene los niveles ciclados neto (niveles totales + niveles de ciclos cancelados) para cada recetario.
    Args:
        fecha: Fecha en formato "YYYY-MM-DD"
    Returns:
        Diccionario {id_recetario: niveles_ciclados_neto}
    """
    session = SessionLocal()
    try:
        query = text("""
            SELECT rxc.id_recetario, SUM(rxc.cantidadNivelesSeleccionados) as niveles_ciclados_neto
            FROM recetarioxciclo rxc
            JOIN ciclodesmoldeo cd ON rxc.id_ciclo_desmoldeo = cd.id
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            GROUP BY rxc.id_recetario
        """)
        result = session.execute(query, {"fecha": fecha})
        # Crear un diccionario {id_recetario: niveles_ciclados_neto}
        return {row[0]: row[1] for row in result.fetchall()}
    finally:
        session.close()

def calcular_eficiencia_bruta_por_recetario(niveles_fallados, niveles_desmoldados):
    """
    Calcula la eficiencia bruta para cada recetario usando la fórmula:
    (1 - niveles_fallados / (niveles_desmoldados + niveles_fallados)) * 100
    Args:
        niveles_fallados: Diccionario {id_recetario: niveles_fallados}
        niveles_desmoldados: Diccionario {id_recetario: niveles_desmoldados}
    Returns:
        Diccionario {id_recetario: eficiencia_bruta}
    """
    eficiencias = {}
    
    # Obtener todos los id_recetario únicos de ambos diccionarios
    todos_ids = set(niveles_fallados.keys()) | set(niveles_desmoldados.keys())
    
    for id_recetario in todos_ids:
        fallados = niveles_fallados.get(id_recetario, 0)
        desmoldados = niveles_desmoldados.get(id_recetario, 0)
        
        # Niveles ciclados neto = niveles desmoldados correctamente + niveles fallados
        niveles_ciclados_neto = desmoldados + fallados
        
        if niveles_ciclados_neto > 0:
            eficiencia = (1 - fallados / niveles_ciclados_neto) * 100
            eficiencias[id_recetario] = round(eficiencia, 2)
        else:
            eficiencias[id_recetario] = 0.0
    
    return eficiencias

def calcular_eficiencia_por_receta(niveles_desmoldados, cantidad_niveles_receta):
    """
    Calcula la eficiencia por receta para cada recetario usando la fórmula:
    (niveles_desmoldados / niveles_receta)
    Args:
        niveles_desmoldados: Diccionario {id_recetario: niveles_desmoldados}
        cantidad_niveles_receta: Diccionario {id_recetario: cantidadNiveles}
    Returns:
        Diccionario {id_recetario: eficiencia_por_receta}
    """
    eficiencias_receta = {}
    
    # Obtener todos los id_recetario únicos de ambos diccionarios
    todos_ids = set(niveles_desmoldados.keys()) | set(cantidad_niveles_receta.keys())
    
    for id_recetario in todos_ids:
        desmoldados = niveles_desmoldados.get(id_recetario, 0)
        niveles_receta = cantidad_niveles_receta.get(id_recetario, 1)  # Evitar división por cero
        
        if niveles_receta > 0:
            eficiencia = desmoldados / niveles_receta
            eficiencias_receta[id_recetario] = round(eficiencia, 2)
        else:
            eficiencias_receta[id_recetario] = 0.0
    
    return eficiencias_receta

def obtener_detalles_ciclos_por_fecha(fecha):
    """
    Obtiene los detalles de todos los ciclos para una fecha específica.
    Args:
        fecha: Fecha en formato "YYYY-MM-DD"
    Returns:
        Lista de tuplas con los detalles de los ciclos
    """
    session = SessionLocal()
    try:
        query = text("""
            SELECT 
                cd.id AS id_ciclo, 
                r.codigoProducto AS producto,
                t.ActualizarTAG AS torre,
                rxc.cantidadNivelesFinalizado AS niveles_desmoldados,
                rxc.cantidadNivelesSeleccionados AS niveles_seleccionados,
                cd.pesoDesmoldado AS peso_desmoldado,
                cd.estadoMaquina AS tipo_fin,
                cd.bandaDesmolde AS cinta_desmolde,
                cd.fecha_inicio AS inicio,
                cd.fecha_fin AS fin,
                cd.tiempoPausado AS tiempo_pausado,
                cd.tiempoDesmolde AS tiempo_total
            FROM ciclodesmoldeo cd
            JOIN recetarioxciclo rxc ON cd.id = rxc.id_ciclo_desmoldeo
            JOIN recetario r ON rxc.id_recetario = r.id
            LEFT JOIN torre t ON 
                t.NTorre = CASE rxc.id_recetario
                    WHEN 1 THEN cd.id_torre
                    WHEN 2 THEN 
                        CASE 
                            WHEN cd.id_torre <= 19 THEN cd.id_torre + 13
                            ELSE cd.id_torre + 90
                        END
                    WHEN 3 THEN cd.id_torre + 32
                    WHEN 4 THEN cd.id_torre + 36
                    WHEN 5 THEN cd.id_torre + 66
                    WHEN 6 THEN cd.id_torre + 96
                    WHEN 7 THEN cd.id_torre + 100
                    WHEN 8 THEN cd.id_torre + 102
                    ELSE cd.id_torre
                END
                AND t.id_recetario = rxc.id_recetario
            WHERE DATE(cd.fecha_fin) = :fecha
            AND cd.fecha_fin IS NOT NULL
            ORDER BY cd.fecha_inicio ASC
        """)
        result = session.execute(query, {"fecha": fecha})
        return result.fetchall()
    finally:
        session.close()

def export_ciclodesmoldeo_efa_to_excel(file_path, fecha_hoy):
    """
    Genera un Excel simplificado para EFA que solo contiene:
    - Hoja 1: RESUMEN DE PRODUCTIVIDAD | EFA ALIMENTOS
    - Hoja 2: RESUMEN DE PRODUCTIVIDAD POR TORRE | EFA ALIMENTOS
    """
    logger.info(f"Exportando datos de ciclodesmoldeo EFA para la fecha: {fecha_hoy}")

    fecha_inicio = pd.to_datetime(fecha_hoy)
    fecha_inicio_str = fecha_inicio.strftime("%Y-%m-%d")

    fecha_fin = pd.to_datetime(fecha_hoy)
    fecha_fin_str = fecha_fin.strftime("%Y-%m-%d")
    
    # Verificar si hay datos antes de crear el Excel
    id_recetarios = obtener_id_recetario_por_fecha(fecha_inicio_str)
    
    # Si no hay datos, retornar False para indicar que no hay registros
    if not id_recetarios:
        logger.info(f"No se encontraron datos para la fecha: {fecha_hoy}")
        return False

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Productividad | EFA"

        logo_path = os.path.join(os.path.dirname(__file__), "static", "cremonarecort.png")
        
        # ------------------ PRIMERA HOJA: RESUMEN DE PRODUCTIVIDAD CLIENTE ------------------
        
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 31.5
            img.width = 126
            ws.add_image(img, "G3")

        ws.merge_cells("A1:G1")
        ws["A1"] = "RESUMEN DE PRODUCTIVIDAD | EFA ALIMENTOS"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws["A3"] = "Fecha inicial de filtrado:"
        ws["A3"].font = Font(size=12, bold=True)
        ws["A4"] = "Fecha final de filtrado:"
        ws["A4"].font = Font(size=12, bold=True)
        ws["B3"] = fecha_inicio_str
        ws["B4"] = fecha_fin_str

        # Obtener datos necesarios
        codigos_producto = obtener_codigos_producto_por_ids_recetario(id_recetarios)
        cantidad_niveles_receta = obtener_cantidad_niveles_receta_por_ids_recetario(id_recetarios)
        cantidad_ciclos = obtener_cantidad_ciclos_por_recetario(fecha_inicio_str)
        ciclos_cancelados = obtener_ciclos_cancelados_por_recetario(fecha_inicio_str)
        pesos_totales = obtener_peso_total_por_recetario(fecha_inicio_str)
        tiempos_totales = obtener_tiempo_total_por_recetario(fecha_inicio_str)
        niveles_desmoldados = obtener_niveles_desmoldados_por_recetario(fecha_inicio_str)
        
        # Obtener datos para calcular eficiencia bruta
        niveles_fallados = obtener_niveles_fallados_por_recetario(fecha_inicio_str)
        niveles_ciclados_neto = obtener_niveles_ciclados_neto_por_recetario(fecha_inicio_str)
        
        # Calcular segundos por nivel, eficiencia bruta y eficiencia por receta
        segundos_por_nivel = obtener_segundos_por_nivel_por_recetario(tiempos_totales, niveles_desmoldados)
        eficiencia_bruta = calcular_eficiencia_bruta_por_recetario(niveles_fallados, niveles_desmoldados)
        eficiencia_por_receta = calcular_eficiencia_por_receta(niveles_desmoldados, cantidad_niveles_receta)

        # Encabezados simplificados
        headers_cliente = [
            "Producto",
            "Cantidad de ciclos",
            "Peso total desmoldado [kg]",
            "Tiempo util desmoldado [HH:MM:SS]",
            "Niveles desmoldados\ncorrectamente",
            "Segundos/Nivel [seg]",
            "Torres equivalentes" #Eficiencia por receta
        ]
        
        # Añadir línea vacía y encabezados
        ws.append([])  # Línea vacía
        ws.append(headers_cliente)

        # Guardar la fila donde empiezan los encabezados para la definición de la tabla
        first_table_first_row = ws.max_row
        
        # Aplicar formato multilínea a los encabezados
        for col in range(1, len(headers_cliente) + 1):
            cell = ws.cell(row=first_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura de la fila de encabezados
        ws.row_dimensions[first_table_first_row].height = 45
        
        if id_recetarios:
            # Ordenar id_recetarios de forma ascendente
            id_recetarios_ordenados = sorted(id_recetarios)
            for id_recetario in id_recetarios_ordenados:
                codigo_producto = codigos_producto.get(id_recetario, "DESCONOCIDO")
                ciclos = cantidad_ciclos.get(id_recetario, 0)
                peso_total = pesos_totales.get(id_recetario, 0.0)
                tiempo_total = tiempos_totales.get(id_recetario, "00:00:00")
                niveles = niveles_desmoldados.get(id_recetario, 0)
                seg_por_nivel = segundos_por_nivel.get(id_recetario, "0")
                eficiencia_receta = eficiencia_por_receta.get(id_recetario, 0.0)
                
                ws.append([
                    codigo_producto,    # Producto
                    ciclos,             # Cantidad de ciclos
                    peso_total,         # Peso total desmoldado
                    tiempo_total,       # Tiempo total desmoldado
                    niveles,            # Niveles desmoldados correctamente
                    seg_por_nivel,      # Segundos/Nivel
                    eficiencia_receta   # Eficiencia por receta
                ])
        else:
            # Si no hay datos, agregar una fila con "Sin datos"
            ws.append([
                "DESCONOCIDO",      # Producto
                0,                  # Cantidad de ciclos
                "0.0",              # Peso total desmoldado
                "00:00:00",         # Tiempo total desmoldado
                0,                  # Niveles desmoldados correctamente
                "0",                # Segundos/Nivel
                0.0                 # Eficiencia por receta
            ])
        
        # Calcular y agregar fila de totales
        total_ciclos = sum(cantidad_ciclos.values())
        total_peso = sum(pesos_totales.values())
        
        # Calcular el tiempo total en segundos
        total_segundos = 0
        for tiempo in tiempos_totales.values():
            partes = tiempo.split(":")
            segundos = int(partes[0]) * 3600 + int(partes[1]) * 60 + int(partes[2])
            total_segundos += segundos
            
        # Convertir segundos totales a formato hh:mm:ss
        horas_total = total_segundos // 3600
        minutos_total = (total_segundos % 3600) // 60
        segundos_total = total_segundos % 60
        tiempo_total_formato = f"{horas_total:02d}:{minutos_total:02d}:{segundos_total:02d}"
        
        # Calcular el total de niveles desmoldados
        total_niveles = sum(niveles_desmoldados.values())
        
        # Calcular segundos por nivel total
        segundos_por_nivel_total = "0"
        if total_niveles > 0:
            segundos_por_nivel_total = int(total_segundos / total_niveles)
        
        # Calcular la eficiencia bruta total usando la fórmula
        total_niveles_fallados = sum(niveles_fallados.values())
        # Niveles ciclados neto = niveles desmoldados correctamente + niveles fallados
        total_niveles_ciclados_neto = total_niveles + total_niveles_fallados
        eficiencia_bruta_total = 0.0
        if total_niveles_ciclados_neto > 0:
            eficiencia_bruta_total = round((1 - (total_niveles_fallados / total_niveles_ciclados_neto)) * 100, 2)
        
        # Calcular eficiencia por receta total (suma de todas las eficiencias individuales)
        eficiencia_por_receta_total = sum(eficiencia_por_receta.values())
        
        # Agregar fila de totales
        ws.append([
            "TOTALES",                    # Producto
            total_ciclos,                 # Cantidad de ciclos
            total_peso,                   # Peso total desmoldado
            tiempo_total_formato,         # Tiempo total desmoldado
            total_niveles,                # Niveles desmoldados correctamente
            segundos_por_nivel_total,     # Segundos/Nivel
            round(eficiencia_por_receta_total, 1)  # Eficiencia por receta total
        ])
        
        # Dar formato a la fila de totales
        fila_totales = ws.max_row
        ws.row_dimensions[fila_totales].height = 30
        
        # Aplicar estilo negrita a la fila de totales
        for col in range(1, 8):  # 7 columnas (A-G)
            cell = ws.cell(row=fila_totales, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center')
            
            # Añadir un borde superior para destacar que es una fila de totales
            thin_border = Border(top=Side(style='thin'))
            cell.border = thin_border
        
        # Crear la primera tabla
        first_table_last_row = ws.max_row
        
        # Agregar la primera tabla
        first_table = Table(displayName="ResumenProductividadCliente", ref=f"A{first_table_first_row}:G{first_table_last_row}")
        first_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        first_table.tableStyleInfo = first_style
        ws.add_table(first_table)
        
        # ------------------ SEGUNDA HOJA: RESUMEN DE PRODUCTIVIDAD POR TORRE CLIENTE ------------------
        
        # Crear una nueva hoja
        ws_torre = wb.create_sheet("Torres | EFA")
        
        # Insertar logo en la nueva hoja
        if os.path.exists(logo_path):
            img2 = XLImage(logo_path)
            img2.height = 31.5
            img2.width = 126
            ws_torre.add_image(img2, "L3")

        # Encabezado de la nueva hoja
        ws_torre.merge_cells("A1:M1")  # Fusionar celdas para el título (13 columnas)
        ws_torre["A1"] = "RESUMEN DE PRODUCTIVIDAD POR TORRE | EFA ALIMENTOS"
        ws_torre["A1"].font = Font(size=16, bold=True)
        ws_torre["A1"].alignment = Alignment(horizontal="center")
        ws_torre["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws_torre["A3"] = "Fecha inicial de filtrado:"
        ws_torre["A3"].font = Font(size=12, bold=True)
        ws_torre["A4"] = "Fecha final de filtrado:"
        ws_torre["A4"].font = Font(size=12, bold=True)
        ws_torre["B3"] = fecha_inicio_str
        ws_torre["B4"] = fecha_fin_str

        # Encabezados para la tabla por torre (sin "Recuento de Fallas")
        headers_torre_cliente = [
            "ID Ciclo",
            "Producto",
            "Torre",
            "Niveles\nDesmoldados",
            "Niveles\nSeleccionados",
            "Peso Desmoldado [kg]",
            "Tipo de Fin",
            "Cinta de\nDesmolde",
            "Inicio [AAAA-MM-DD HH:MM:SS]",
            "Fin [AAAA-MM-DD HH:MM:SS]",
            "Tiempo Desmolde\n[MM:SS]",
            "Tiempo Pausado\n[MM:SS]",
            "Tiempo Útil\n[MM:SS]"
        ]
        
        # Añadir línea vacía y encabezados
        ws_torre.append([])  # Línea vacía
        ws_torre.append(headers_torre_cliente)
        
        # Guardar la fila donde empiezan los encabezados para la definición de la tabla
        torre_table_first_row = ws_torre.max_row
        
        # Aplicar formato multilínea a los encabezados de torres
        for col in range(1, len(headers_torre_cliente) + 1):
            cell = ws_torre.cell(row=torre_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura de la fila de encabezados de torres
        ws_torre.row_dimensions[torre_table_first_row].height = 45
        
        # Obtener datos detallados de los ciclos
        detalles_ciclos = obtener_detalles_ciclos_por_fecha(fecha_inicio_str)
        
        if detalles_ciclos:
            for ciclo in detalles_ciclos:
                # Formatear fechas
                fecha_inicio = ciclo[8].strftime("%Y-%m-%d %H:%M:%S") if ciclo[8] else "N/A"
                fecha_fin = ciclo[9].strftime("%Y-%m-%d %H:%M:%S") if ciclo[9] else "N/A"
                
                # Obtener tiempos directamente de la base
                tiempo_pausado_str = ciclo[10] or "00:00"
                tiempo_desmolde_str = ciclo[11] or "00:00"
                
                # Convertir a formato MM:SS
                tiempo_pausado_mm_ss = tiempo_a_mm_ss(tiempo_pausado_str)
                tiempo_desmolde_mm_ss = tiempo_a_mm_ss(tiempo_desmolde_str)
                tiempo_util_mm_ss = calcular_tiempo_util_mm_ss(tiempo_desmolde_str, tiempo_pausado_str)
                
                # Obtener tipo de fin
                tipo_fin = ciclo[6] or "N/A"
                
                ws_torre.append([
                    ciclo[0],                  # ID Ciclo
                    ciclo[1] or "DESCONOCIDO", # Producto
                    ciclo[2] or "N/A",         # Torre
                    ciclo[3] or 0,             # Niveles Desmoldados
                    ciclo[4] or 0,             # Niveles Seleccionados
                    ciclo[5] or 0, # Peso Desmoldado
                    tipo_fin,                  # Tipo de Fin
                    ciclo[7] or "N/A",         # Cinta de Desmolde
                    fecha_inicio,              # Inicio
                    fecha_fin,                 # Fin
                    tiempo_desmolde_mm_ss,     # Tiempo Desmolde [MM:SS]
                    tiempo_pausado_mm_ss,      # Tiempo Pausado [MM:SS]
                    tiempo_util_mm_ss          # Tiempo Útil [MM:SS]
                ])
        else:
            # Si no hay datos, agregar una fila con "Sin datos"
            ws_torre.append([
                "Sin datos",    # ID Ciclo
                "DESCONOCIDO",  # Producto
                "N/A",          # Torre
                0,              # Niveles Desmoldados
                0,              # Niveles Seleccionados
                0,      # Peso Desmoldado
                "N/A",          # Tipo de Fin
                "N/A",          # Cinta de Desmolde
                "N/A",          # Inicio
                "N/A",          # Fin
                "00:00",        # Tiempo Desmolde [MM:SS]
                "00:00",        # Tiempo Pausado [MM:SS]
                "00:00"         # Tiempo Útil [MM:SS]
            ])
        
        # Definir la tabla de torre
        torre_table_last_row = ws_torre.max_row
        
        # Aplicar formato de color rojo claro a las filas con "CANCELADO AL INICIAR"
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Recorrer las filas de datos (excluyendo encabezados)
        for row_num in range(torre_table_first_row + 1, torre_table_last_row + 1):
            tipo_fin_cell = ws_torre[f"G{row_num}"]  # Columna G es "Tipo de Fin"
            if tipo_fin_cell.value == "CANCELADO AL INICIAR":
                # Aplicar color de fondo rojo claro a toda la fila
                for col_num in range(1, 14):  # Columnas A-M (13 columnas)
                    cell = ws_torre.cell(row=row_num, column=col_num)
                    cell.fill = light_red_fill
        
        # Agregar la tabla de torre
        torre_table = Table(displayName="ResumenProductividadTorreCliente", ref=f"A{torre_table_first_row}:M{torre_table_last_row}")
        torre_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        torre_table.tableStyleInfo = torre_style
        ws_torre.add_table(torre_table)
        
        # Ajustar ancho de columnas para ambas hojas
        for sheet in [ws, ws_torre]:
            for col in sheet.columns:
                max_data_length = 0
                max_header_length = 0
                column = None
                is_header_multiline = False
                
                for cell in col:
                    if hasattr(cell, "column_letter"):
                        column = cell.column_letter
                    else:
                        continue
                        
                    try:
                        if cell.value:
                            cell_str = str(cell.value)
                            
                            # Detectar si es un encabezado (filas 5-7 típicamente contienen encabezados)
                            if 5 <= cell.row <= 8 and ('[' in cell_str or 'Tiempo' in cell_str or 'Nivel' in cell_str or 'Peso' in cell_str):
                                # Es un encabezado
                                if '\n' in cell_str:
                                    lines = cell_str.split('\n')
                                    max_header_length = max(len(line) for line in lines)
                                    is_header_multiline = True
                                else:
                                    max_header_length = len(cell_str)
                            else:
                                # Es dato normal
                                if cell.row > 8:  # Evitar títulos y fechas
                                    max_data_length = max(max_data_length, len(cell_str))
                    except:
                        pass
                
                if column:
                    # Establecer ancho fijo de 25 para la primera columna (A)
                    if column == "A":
                        final_width = 25
                    else:
                        # Calcular ancho basándose en el contenido más largo (header vs datos)
                        content_width = max(max_header_length, max_data_length)
                        
                        # Establecer rangos específicos por tipo de columna
                        if column == "B":
                            # Para columna Producto, usar el mayor entre header y contenido, con mínimo 10
                            final_width = min(max(content_width + 2, 10), 25)
                        elif "Torres equivalentes" in str(sheet[column + "6"].value or "") or any("Torres equivalentes" in str(cell.value or "") for cell in sheet[column]):
                            final_width = 19  # Ancho específico para "Torres equivalentes"
                        elif "Tiempo" in str(max_header_length) or "MM:SS" in str(max_header_length):
                            final_width = 12  # Columnas de tiempo
                        elif is_header_multiline:
                            # Para encabezados multilínea, usar la línea más larga + pequeño margen
                            final_width = min(max(max_header_length + 2, 10), 20)
                        elif "[kg]" in str(sheet[column + "6"].value or "") or "Peso" in str(sheet[column + "6"].value or ""):
                            final_width = min(max(content_width + 2, 12), 16)  # Columnas de peso
                        elif "[seg]" in str(sheet[column + "6"].value or ""):
                            final_width = 21  # Segundos/Nivel
                        else:
                            final_width = min(max(content_width + 2, 10), 25)  # Otras columnas
                    
                    sheet.column_dimensions[column].width = final_width

        wb.save(file_path)
        logger.info(f"Archivo Excel EFA generado en: {file_path}")
        return True  # Retornar True indicando que se generó el Excel exitosamente
    except Exception as e:
        logger.error(f"Error exportando a Excel EFA: {e}")
        raise

def export_ciclodesmoldeo_to_excel(file_path, fecha_hoy):
    logger.info(f"Exportando datos de ciclodesmoldeo para la fecha: {fecha_hoy}")

    fecha_inicio = pd.to_datetime(fecha_hoy)
    fecha_inicio_str = fecha_inicio.strftime("%Y-%m-%d")

    fecha_fin = pd.to_datetime(fecha_hoy)
    fecha_fin_str = fecha_fin.strftime("%Y-%m-%d")
    
    # Verificar si hay datos antes de crear el Excel
    id_recetarios = obtener_id_recetario_por_fecha(fecha_inicio_str)
    
    # Si no hay datos, retornar False para indicar que no hay registros
    if not id_recetarios:
        logger.info(f"No se encontraron datos para la fecha: {fecha_hoy}")
        return False

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Productividad | INGENIERIA"

        logo_path = os.path.join(os.path.dirname(__file__), "static", "cremonarecort.png")
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 31.5
            img.width = 126
            ws.add_image(img, "K3")

        ws.merge_cells("A1:K1")
        ws["A1"] = "RESUMEN DE PRODUCTIVIDAD | CREMINOX"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws["A3"] = "Fecha inicial de filtrado:"
        ws["A3"].font = Font(size=12, bold=True)
        ws["A4"] = "Fecha final de filtrado:"
        ws["A4"].font = Font(size=12, bold=True)
        ws["B3"] = fecha_inicio_str
        ws["B4"] = fecha_fin_str

        codigos_producto = obtener_codigos_producto_por_ids_recetario(id_recetarios)
        cantidad_niveles_receta = obtener_cantidad_niveles_receta_por_ids_recetario(id_recetarios)
        cantidad_ciclos = obtener_cantidad_ciclos_por_recetario(fecha_inicio_str)
        ciclos_cancelados = obtener_ciclos_cancelados_por_recetario(fecha_inicio_str)
        pesos_totales = obtener_peso_total_por_recetario(fecha_inicio_str)
        tiempos_totales = obtener_tiempo_total_por_recetario(fecha_inicio_str)
        niveles_desmoldados = obtener_niveles_desmoldados_por_recetario(fecha_inicio_str)
        
        # Obtener datos para calcular eficiencia bruta
        niveles_fallados = obtener_niveles_fallados_por_recetario(fecha_inicio_str)
        niveles_ciclados_neto = obtener_niveles_ciclados_neto_por_recetario(fecha_inicio_str)
        
        # Calcular segundos por nivel, porcentaje de fallas, eficiencia bruta y eficiencia por receta
        segundos_por_nivel = obtener_segundos_por_nivel_por_recetario(tiempos_totales, niveles_desmoldados)
        porcentaje_fallas = calcular_porcentaje_fallas(cantidad_ciclos, ciclos_cancelados)
        eficiencia_bruta = calcular_eficiencia_bruta_por_recetario(niveles_fallados, niveles_desmoldados)
        eficiencia_por_receta = calcular_eficiencia_por_receta(niveles_desmoldados, cantidad_niveles_receta)

        # Añadir las nuevas columnas a los encabezados
        headers = [
            "ID Receta",
            "Producto",
            "Cantidad de ciclos",
            "Peso total desmoldado [kg]",
            "Tiempo util desmoldado [HH:MM:SS]",
            "Niveles desmoldados\ncorrectamente",
            "Segundos/Nivel [seg]",
            "Eficiencia bruta",
            "% Falla",
            "Torres equivalentes",
            "Kilos por hora\nproducidos [kg/h]"
        ]
        ws.append([])  # Línea vacía
        ws.append(headers)

        # Aplicar formato multilínea a los encabezados principales
        header_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura de la fila de encabezados principales
        ws.row_dimensions[header_row].height = 45

        if id_recetarios:
            # Ordenar id_recetarios de forma ascendente
            id_recetarios_ordenados = sorted(id_recetarios)
            for id_recetario in id_recetarios_ordenados:
                codigo_producto = codigos_producto.get(id_recetario, "DESCONOCIDO")
                ciclos = cantidad_ciclos.get(id_recetario, 0)
                peso_total = pesos_totales.get(id_recetario, 0.0)
                tiempo_total = tiempos_totales.get(id_recetario, "00:00:00")
                niveles = niveles_desmoldados.get(id_recetario, 0)
                seg_por_nivel = segundos_por_nivel.get(id_recetario, "0")
                falla_porcentaje = porcentaje_fallas.get(id_recetario, 0)
                eficiencia = eficiencia_bruta.get(id_recetario, 0.0)
                eficiencia_receta = eficiencia_por_receta.get(id_recetario, 0.0)
                
                # Calcular kilos por hora
                kilos_por_hora = 0.0
                if tiempo_total != "00:00:00":
                    # Convertir tiempo HH:MM:SS a horas decimales
                    partes_tiempo = tiempo_total.split(":")
                    horas_decimales = int(partes_tiempo[0]) + int(partes_tiempo[1])/60 + int(partes_tiempo[2])/3600
                    if horas_decimales > 0:
                        kilos_por_hora = round(peso_total / horas_decimales, 2)
                
                # Añadir las columnas en la fila
                ws.append([
                    str(id_recetario),
                    codigo_producto,
                    ciclos,
                    peso_total,
                    tiempo_total,
                    niveles,
                    seg_por_nivel,
                    f"{eficiencia}%",       # Eficiencia bruta con formato de porcentaje
                    f"{falla_porcentaje}%", # % Falla
                    eficiencia_receta,       # Eficiencia por receta
                    kilos_por_hora             # Kilos por hora
                ])
        else:
            # Si no hay datos, agregar una fila con "Sin datos"
            ws.append([
                "Sin datos",
                "DESCONOCIDO",
                0,
                "0.0",
                "00:00",
                0,
                "0",
                "0.0%",             # Eficiencia bruta
                "0%",               # % Falla
                0.0,                # Eficiencia por receta
                0.0                 # Kilos por hora
            ])
            
        # Agregar la fila de totales
        # Calcular los totales
        total_ciclos = sum(cantidad_ciclos.values())
        total_peso = sum(pesos_totales.values())
        total_ciclos_cancelados = sum(ciclos_cancelados.values())
        
        # Calcular el tiempo total en segundos
        total_segundos = 0
        for tiempo in tiempos_totales.values():
            partes = tiempo.split(":")
            segundos = int(partes[0]) * 3600 + int(partes[1]) * 60 + int(partes[2])
            total_segundos += segundos
            
        # Convertir segundos totales a formato hh:mm:ss
        horas_total = total_segundos // 3600
        minutos_total = (total_segundos % 3600) // 60
        segundos_total = total_segundos % 60
        tiempo_total_formato = f"{horas_total:02d}:{minutos_total:02d}:{segundos_total:02d}"
        
        # Calcular el total de niveles desmoldados
        total_niveles = sum(niveles_desmoldados.values())
        
        # Calcular segundos por nivel total
        segundos_por_nivel_total = "0"
        if total_niveles > 0:
            segundos_por_nivel_total = int(total_segundos / total_niveles)
        
        # Calcular el porcentaje de fallas total (ciclos cancelados totales / ciclos totales * 100)
        porcentaje_falla_total = 0
        if total_ciclos > 0:
            porcentaje_falla_total = round((total_ciclos_cancelados / total_ciclos) * 100, 2)
        
        # Calcular la eficiencia bruta total usando la fórmula
        total_niveles_fallados = sum(niveles_fallados.values())
        # Niveles ciclados neto = niveles desmoldados correctamente + niveles fallados
        total_niveles_ciclados_neto = total_niveles + total_niveles_fallados
        eficiencia_bruta_total = 0.0
        if total_niveles_ciclados_neto > 0:
            eficiencia_bruta_total = round((1 - (total_niveles_fallados / total_niveles_ciclados_neto)) * 100, 2)
        
        # Calcular eficiencia por receta total (suma de todas las eficiencias individuales)
        eficiencia_por_receta_total = sum(eficiencia_por_receta.values())
        
        # Calcular kilos por hora total
        kilos_por_hora_total = 0.0
        if total_segundos > 0:
            horas_totales_decimales = total_segundos / 3600
            kilos_por_hora_total = round(total_peso / horas_totales_decimales, 2)
        
        # Agregar la fila de totales
        ws.append([
            "TOTALES",                        # ID Receta
            "",                               # Producto
            total_ciclos,                     # Cantidad de ciclos
            total_peso,                       # Peso total desmoldado
            tiempo_total_formato,             # Tiempo total desmoldado
            total_niveles,                    # Niveles desmoldados correctamente
            segundos_por_nivel_total,         # Segundos/Nivel
            f"{eficiencia_bruta_total}%",     # Eficiencia bruta total
            f"{porcentaje_falla_total}%",     # % Falla (total)
            round(eficiencia_por_receta_total, 1),  # Eficiencia por receta total
            kilos_por_hora_total                # Kilos por hora total
        ])
        
        # Dar formato a la fila de totales
        fila_totales = ws.max_row
        ws.row_dimensions[fila_totales].height = 30  # El doble de la altura normal
        
        # Aplicar estilo negrita a la fila de totales
        for col in range(1, 12):  # 11 columnas (A-K)
            cell = ws.cell(row=fila_totales, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center')
            
            # Añadir un borde superior para destacar que es una fila de totales
            thin_border = Border(top=Side(style='thin'))
            cell.border = thin_border

        first_table_row = ws.max_row - len(id_recetarios or [1]) - 1  # La fila de encabezados
        last_table_row = ws.max_row  # La última fila de datos (incluyendo totales)

        # Agregar la primera tabla
        table = Table(displayName="ResumenProductividad", ref=f"A{first_table_row}:K{last_table_row}")
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Ajustar ancho de la columna B basándose SOLO en la primera tabla
        max_producto_length = 0
        for row in range(first_table_row + 1, last_table_row + 1):  # +1 para saltar encabezados
            cell_value = ws[f"B{row}"].value
            if cell_value and str(cell_value) != "TOTALES":
                max_producto_length = max(max_producto_length, len(str(cell_value)))
        
        # Establecer ancho para columna B basándose solo en la primera tabla
        first_table_product_width = min(max(max_producto_length + 2, 10), 25)
        ws.column_dimensions["B"].width = first_table_product_width

        # ------------------ SEGUNDA TABLA: RESUMEN DE PRODUCTIVIDAD CLIENTE ------------------
        
        # Guardar la última fila de la primera tabla
        ultima_fila_primera_tabla = ws.max_row
        
        # Crear 2 filas vacías de separación explícitas (en lugar de usar append)
        for i in range(1, 3):
            nueva_fila = ultima_fila_primera_tabla + i
            for col in range(1, 10):  # Asegurarse de que todas las columnas estén vacías
                ws.cell(row=nueva_fila, column=col, value=None)
        
        # El título de la segunda tabla debe estar 2 filas después de la última fila de la primera tabla
        segunda_tabla_titulo_fila = ultima_fila_primera_tabla + 3
        
        ws.merge_cells(f"A{segunda_tabla_titulo_fila}:G{segunda_tabla_titulo_fila}")
        cell = ws.cell(row=segunda_tabla_titulo_fila, column=1)
        cell.value = "RESUMEN DE PRODUCTIVIDAD | EFA ALIMENTOS"
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        fecha_row = segunda_tabla_titulo_fila + 2  # Una fila vacía después del título
        subtitulo_row = segunda_tabla_titulo_fila + 3  # Una fila vacía después del título
        
        # Añadir la fecha
        ws.cell(row=fecha_row, column=1, value="Fecha inicial de filtrado:")
        ws.cell(row=fecha_row, column=1).font = Font(size=12, bold=True)
        ws.cell(row=fecha_row, column=2, value=fecha_inicio_str)

        ws.cell(row=subtitulo_row, column=1, value="Fecha final de filtrado:")
        ws.cell(row=subtitulo_row, column=1).font = Font(size=12, bold=True)
        ws.cell(row=subtitulo_row, column=2, value=fecha_fin_str)

        empty_row = subtitulo_row + 1
        
        # Insertar el segundo logo
        if os.path.exists(logo_path):
            img2 = XLImage(logo_path)
            img2.height = 31.5
            img2.width = 126
            ws.add_image(img2, f"G{fecha_row}")  # Logo alineado a la derecha

        # Añadir los encabezados simplificados (después de la fecha)
        headers_row = empty_row + 1  # Una fila vacía después de la fecha/logo
        
        headers_cliente = [
            "Producto",
            "Cantidad de ciclos",
            "Peso total desmoldado [kg]",
            "Tiempo util desmoldado [HH:MM:SS]",
            "Niveles desmoldados\ncorrectamente",
            "Segundos/Nivel [seg]",
            "Torres equivalentes" #Eficiencia por receta
        ]
        
        # Agregar encabezados
        for col, header in enumerate(headers_cliente, 1):
            ws.cell(row=headers_row, column=col, value=header)
        
        # Aplicar formato multilínea a los encabezados de cliente
        for col in range(1, len(headers_cliente) + 1):
            cell = ws.cell(row=headers_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura de la fila de encabezados de cliente
        ws.row_dimensions[headers_row].height = 45
                
        # Guardar la fila donde empiezan los encabezados para la definición de la tabla
        second_table_first_row = headers_row
        
        if id_recetarios:
            # Ordenar id_recetarios de forma ascendente
            id_recetarios_ordenados = sorted(id_recetarios)
            for id_recetario in id_recetarios_ordenados:
                codigo_producto = codigos_producto.get(id_recetario, "DESCONOCIDO")
                ciclos = cantidad_ciclos.get(id_recetario, 0)
                peso_total = pesos_totales.get(id_recetario, 0.0)
                tiempo_total = tiempos_totales.get(id_recetario, "00:00:00")
                niveles = niveles_desmoldados.get(id_recetario, 0)
                seg_por_nivel = segundos_por_nivel.get(id_recetario, "0")
                eficiencia_receta = eficiencia_por_receta.get(id_recetario, 0.0)
                
                # Añadir solo las columnas especificadas en la segunda tabla
                ws.append([
                    codigo_producto,    # Producto
                    ciclos,             # Cantidad de ciclos
                    peso_total,         # Peso total desmoldado
                    tiempo_total,       # Tiempo total desmoldado
                    niveles,            # Niveles desmoldados correctamente
                    seg_por_nivel,      # Segundos/Nivel
                    eficiencia_receta   # Torres equivalentes (Eficiencia por receta)
                ])
        else:
            # Si no hay datos, agregar una fila con "Sin datos"
            ws.append([
                "DESCONOCIDO",      # Producto
                0,                  # Cantidad de ciclos
                "0.0",              # Peso total desmoldado
                "00:00:00",         # Tiempo total desmoldado
                0,                  # Niveles desmoldados correctamente
                "0",                # Segundos/Nivel
                0.0                 # Torres equivalentes (Eficiencia por receta)
            ])
        
        # Agregar fila de totales a la segunda tabla
        ws.append([
            "TOTALES",            # Producto
            total_ciclos,         # Cantidad de ciclos
            total_peso,           # Peso total desmoldado
            tiempo_total_formato, # Tiempo total desmoldado
            total_niveles,        # Niveles desmoldados correctamente
            segundos_por_nivel_total,  # Segundos/Nivel
            round(eficiencia_por_receta_total, 1)  # Torres equivalentes (Eficiencia por receta total)
        ])
        
        # Dar formato a la fila de totales de la segunda tabla
        fila_totales_segunda = ws.max_row
        ws.row_dimensions[fila_totales_segunda].height = 30  # El doble de la altura normal
        
        # Aplicar estilo negrita a la fila de totales de la segunda tabla
        for col in range(1, 8):  # 7 columnas (A-G)
            cell = ws.cell(row=fila_totales_segunda, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center')
            
            # Añadir un borde superior para destacar que es una fila de totales
            thin_border = Border(top=Side(style='thin'))
            cell.border = thin_border
        
        # Crear la segunda tabla
        second_table_last_row = ws.max_row  # La última fila de datos (incluyendo totales)
        
        # Agregar la segunda tabla
        second_table = Table(displayName="ResumenProductividadCliente", ref=f"A{second_table_first_row}:G{second_table_last_row}")
        second_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        second_table.tableStyleInfo = second_style
        ws.add_table(second_table)
        
        # ------------------ TERCERA PÁGINA: RESUMEN POR TORRE ------------------
        
        # Crear una nueva hoja
        ws_torre = wb.create_sheet("Torres | INGENIERIA")
        
        # Insertar logo en la nueva hoja
        if os.path.exists(logo_path):
            img3 = XLImage(logo_path)
            img3.height = 31.5
            img3.width = 126
            ws_torre.add_image(img3, "N3")

        # Encabezado de la nueva hoja
        ws_torre.merge_cells("A1:O1")  # Fusionar celdas para el título (15 columnas)
        ws_torre["A1"] = "RESUMEN DE PRODUCTIVIDAD POR TORRE | CREMINOX"
        ws_torre["A1"].font = Font(size=16, bold=True)
        ws_torre["A1"].alignment = Alignment(horizontal="center")
        ws_torre["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws_torre["A3"] = "Fecha inicial de filtrado:"
        ws_torre["A3"].font = Font(size=12, bold=True)
        ws_torre["A4"] = "Fecha final de filtrado:"
        ws_torre["A4"].font = Font(size=12, bold=True)
        ws_torre["B3"] = fecha_inicio_str
        ws_torre["B4"] = fecha_fin_str

        # Encabezados para la tabla por torre
        headers_torre = [
            "ID Ciclo",
            "Producto",
            "Torre",
            "Niveles\nDesmoldados",
            "Niveles\nSeleccionados",
            "Peso Desmoldado [kg]",
            "Tipo de Fin",
            "Cinta de\nDesmolde",
            "Inicio [AAAA-MM-DD HH:MM:SS]",
            "Fin [AAAA-MM-DD HH:MM:SS]",
            "Tiempo Desmolde\n[MM:SS]",
            "Tiempo Pausado\n[MM:SS]",
            "Tiempo Útil\n[MM:SS]",
            "Recuento de Fallas",
            "Tiempo entre ciclos\n[HH:MM:SS]"
        ]
        
        # Añadir línea vacía y encabezados
        ws_torre.append([])  # Línea vacía
        ws_torre.append(headers_torre)
        
        # Guardar la fila donde empiezan los encabezados para la definición de la tabla
        torre_table_first_row = ws_torre.max_row
        
        # Aplicar formato multilínea a los encabezados de torres ingeniería
        for col in range(1, len(headers_torre) + 1):
            cell = ws_torre.cell(row=torre_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura de la fila de encabezados de torres ingeniería
        ws_torre.row_dimensions[torre_table_first_row].height = 45
        
        # Obtener datos detallados de los ciclos
        detalles_ciclos = obtener_detalles_ciclos_por_fecha(fecha_inicio_str)
        
        # Contador de fallas acumulado
        contador_fallas = 0
        fecha_fin_anterior = None  # Para almacenar la fecha de fin del ciclo anterior
        
        if detalles_ciclos:
            for i, ciclo in enumerate(detalles_ciclos):
                # Formatear fechas
                fecha_inicio = ciclo[8].strftime("%Y-%m-%d %H:%M:%S") if ciclo[8] else "N/A"
                fecha_fin = ciclo[9].strftime("%Y-%m-%d %H:%M:%S") if ciclo[9] else "N/A"
                
                # Calcular tiempo entre ciclos
                if i == 0:  # Primer ciclo
                    tiempo_entre_ciclos = "00:00:00"
                else:
                    tiempo_entre_ciclos = calcular_tiempo_entre_ciclos(ciclo[8], fecha_fin_anterior)
                
                # Actualizar fecha_fin_anterior para el siguiente ciclo
                fecha_fin_anterior = ciclo[9]
                
                # Obtener tiempos directamente de la base
                tiempo_pausado_str = ciclo[10] or "00:00"
                tiempo_desmolde_str = ciclo[11] or "00:00"
                
                # Convertir a formato MM:SS
                tiempo_pausado_mm_ss = tiempo_a_mm_ss(tiempo_pausado_str)
                tiempo_desmolde_mm_ss = tiempo_a_mm_ss(tiempo_desmolde_str)
                tiempo_util_mm_ss = calcular_tiempo_util_mm_ss(tiempo_desmolde_str, tiempo_pausado_str)
                
                # Verificar si es una falla y determinar qué mostrar en la columna
                tipo_fin = ciclo[6] or "N/A"
                if tipo_fin in ["CANCELADO", "CANCELADO AL INICIAR"]:
                    contador_fallas += 1
                    valor_fallas = contador_fallas  # Mostrar el número solo cuando hay falla
                else:
                    valor_fallas = "-"  # Mostrar guión cuando no hay falla
                
                ws_torre.append([
                    ciclo[0],                  # ID Ciclo
                    ciclo[1] or "DESCONOCIDO", # Producto
                    ciclo[2] or "N/A",         # Torre
                    ciclo[3] or 0,             # Niveles Desmoldados
                    ciclo[4] or 0,             # Niveles Seleccionados
                    ciclo[5] or 0,             # Peso Desmoldado
                    tipo_fin,                  # Tipo de Fin
                    ciclo[7] or "N/A",         # Cinta de Desmolde
                    fecha_inicio,              # Inicio
                    fecha_fin,                 # Fin
                    tiempo_desmolde_mm_ss,     # Tiempo Desmolde [MM:SS]
                    tiempo_pausado_mm_ss,      # Tiempo Pausado [MM:SS]
                    tiempo_util_mm_ss,         # Tiempo Útil [MM:SS]
                    valor_fallas,              # Recuento de Fallas (número solo en fallas, "-" en el resto)
                    tiempo_entre_ciclos        # Tiempo entre ciclos [MM:SS]
                ])
        else:
            # Si no hay datos, agregar una fila con "Sin datos"
            ws_torre.append([
                "Sin datos",    # ID Ciclo
                "DESCONOCIDO",  # Producto
                "N/A",          # Torre
                0,              # Niveles Desmoldados
                0,              # Niveles Seleccionados
                0,              # Peso Desmoldado
                "N/A",          # Tipo de Fin
                "N/A",          # Cinta de Desmolde
                "N/A",          # Inicio
                "N/A",          # Fin
                "00:00",        # Tiempo Desmolde [MM:SS]
                "00:00",        # Tiempo Pausado [MM:SS]
                "00:00",        # Tiempo Útil [MM:SS]
                "-",            # Recuento de Fallas
                "00:00:00"      # Tiempo entre ciclos [HH:MM:SS]
            ])
        
        # Definir la tabla de torre
        torre_table_last_row = ws_torre.max_row
        
        # Aplicar formato de color rojo claro a las filas con "CANCELADO AL INICIAR"
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Recorrer las filas de datos (excluyendo encabezados)
        for row_num in range(torre_table_first_row + 1, torre_table_last_row + 1):
            tipo_fin_cell = ws_torre[f"G{row_num}"]  # Columna G es "Tipo de Fin"
            if tipo_fin_cell.value == "CANCELADO AL INICIAR":
                # Aplicar color de fondo rojo claro a toda la fila
                for col_num in range(1, 16):  # Columnas A-O (15 columnas)
                    cell = ws_torre.cell(row=row_num, column=col_num)
                    cell.fill = light_red_fill
        
        # Agregar la tabla de torre
        torre_table = Table(displayName="ResumenProductividadTorre", ref=f"A{torre_table_first_row}:O{torre_table_last_row}")
        torre_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        torre_table.tableStyleInfo = torre_style
        ws_torre.add_table(torre_table)
        
        # ------------------ SEGUNDA TABLA: RESUMEN DE PRODUCTIVIDAD POR TORRE CLIENTE ------------------
        
        # Guardar la última fila de la primera tabla de torres
        ultima_fila_primera_tabla_torres = ws_torre.max_row
        
        # Crear 2 filas vacías de separación explícitas
        for i in range(1, 3):
            nueva_fila = ultima_fila_primera_tabla_torres + i
            for col in range(1, 16):  # Asegurarse de que todas las columnas estén vacías (15 columnas)
                ws_torre.cell(row=nueva_fila, column=col, value=None)
        
        # El título de la segunda tabla debe estar 2 filas después de la última fila de la primera tabla
        segunda_tabla_torres_titulo_fila = ultima_fila_primera_tabla_torres + 3
        
        ws_torre.merge_cells(f"A{segunda_tabla_torres_titulo_fila}:M{segunda_tabla_torres_titulo_fila}")
        cell = ws_torre.cell(row=segunda_tabla_torres_titulo_fila, column=1)
        cell.value = "RESUMEN DE PRODUCTIVIDAD POR TORRE | EFA ALIMENTOS"
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        fecha_row_torres = segunda_tabla_torres_titulo_fila + 2  # Una fila vacía después del título
        subtitulo_row_torres = segunda_tabla_torres_titulo_fila + 3  # Una fila vacía después del título
        
        # Añadir la fecha para la segunda tabla de torres
        ws_torre.cell(row=fecha_row_torres, column=1, value="Fecha inicial de filtrado:")
        ws_torre.cell(row=fecha_row_torres, column=1).font = Font(size=12, bold=True)
        ws_torre.cell(row=fecha_row_torres, column=2, value=fecha_inicio_str)

        ws_torre.cell(row=subtitulo_row_torres, column=1, value="Fecha final de filtrado:")
        ws_torre.cell(row=subtitulo_row_torres, column=1).font = Font(size=12, bold=True)
        ws_torre.cell(row=subtitulo_row_torres, column=2, value=fecha_fin_str)

        empty_row_torres = subtitulo_row_torres + 1
        
        # Insertar el logo para la segunda tabla de torres
        if os.path.exists(logo_path):
            img4 = XLImage(logo_path)
            img4.height = 31.5
            img4.width = 126
            ws_torre.add_image(img4, f"L{fecha_row_torres}")  # Logo alineado a la derecha

        # Encabezados para la segunda tabla por torre (sin "Recuento de Fallas")
        headers_torre_cliente = [
            "ID Ciclo",
            "Producto",
            "Torre",
            "Niveles\nDesmoldados",
            "Niveles\nSeleccionados",
            "Peso Desmoldado [kg]",
            "Tipo de Fin",
            "Cinta de\nDesmolde",
            "Inicio [AAAA-MM-DD HH:MM:SS]",
            "Fin [AAAA-MM-DD HH:MM:SS]",
            "Tiempo Desmolde\n[MM:SS]",
            "Tiempo Pausado\n[MM:SS]",
            "Tiempo Útil\n[MM:SS]"
        ]
        
        # Añadir los encabezados de la segunda tabla (después de la fecha)
        headers_row_torres = empty_row_torres + 1  # Una fila vacía después de la fecha/logo
        
        # Agregar encabezados
        for col, header in enumerate(headers_torre_cliente, 1):
            ws_torre.cell(row=headers_row_torres, column=col, value=header)
        
        # Aplicar formato multilínea a los encabezados de la segunda tabla de torres
        for col in range(1, len(headers_torre_cliente) + 1):
            cell = ws_torre.cell(row=headers_row_torres, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura de la fila de encabezados de la segunda tabla de torres
        ws_torre.row_dimensions[headers_row_torres].height = 45
                
        # Guardar la fila donde empiezan los encabezados para la definición de la segunda tabla
        second_torre_table_first_row = headers_row_torres
        
        # Añadir los mismos datos de la primera tabla pero sin la columna "Recuento de Fallas"
        if detalles_ciclos:
            for ciclo in detalles_ciclos:
                # Formatear fechas
                fecha_inicio = ciclo[8].strftime("%Y-%m-%d %H:%M:%S") if ciclo[8] else "N/A"
                fecha_fin = ciclo[9].strftime("%Y-%m-%d %H:%M:%S") if ciclo[9] else "N/A"
                
                # Obtener tiempos directamente de la base
                tiempo_pausado_str = ciclo[10] or "00:00"
                tiempo_desmolde_str = ciclo[11] or "00:00"
                
                # Convertir a formato MM:SS
                tiempo_pausado_mm_ss = tiempo_a_mm_ss(tiempo_pausado_str)
                tiempo_desmolde_mm_ss = tiempo_a_mm_ss(tiempo_desmolde_str)
                tiempo_util_mm_ss = calcular_tiempo_util_mm_ss(tiempo_desmolde_str, tiempo_pausado_str)
                
                # Obtener tipo de fin
                tipo_fin = ciclo[6] or "N/A"
                
                ws_torre.append([
                    ciclo[0],                  # ID Ciclo
                    ciclo[1] or "DESCONOCIDO", # Producto
                    ciclo[2] or "N/A",         # Torre
                    ciclo[3] or 0,             # Niveles Desmoldados
                    ciclo[4] or 0,             # Niveles Seleccionados
                    ciclo[5] or 0, # Peso Desmoldado
                    tipo_fin,                  # Tipo de Fin
                    ciclo[7] or "N/A",         # Cinta de Desmolde
                    fecha_inicio,              # Inicio
                    fecha_fin,                 # Fin
                    tiempo_desmolde_mm_ss,     # Tiempo Desmolde [MM:SS]
                    tiempo_pausado_mm_ss,      # Tiempo Pausado [MM:SS]
                    tiempo_util_mm_ss          # Tiempo Útil [MM:SS]
                    # NO incluir "Recuento de Fallas"
                ])
        else:
            # Si no hay datos, agregar una fila con "Sin datos"
            ws_torre.append([
                "Sin datos",    # ID Ciclo
                "DESCONOCIDO",  # Producto
                "N/A",          # Torre
                0,              # Niveles Desmoldados
                0,              # Niveles Seleccionados
                0,      # Peso Desmoldado
                "N/A",          # Tipo de Fin
                "N/A",          # Cinta de Desmolde
                "N/A",          # Inicio
                "N/A",          # Fin
                "00:00",        # Tiempo Desmolde [MM:SS]
                "00:00",        # Tiempo Pausado [MM:SS]
                "00:00"         # Tiempo Útil [MM:SS]
                # NO incluir "Recuento de Fallas"
            ])
        
        # Crear la segunda tabla de torre
        second_torre_table_last_row = ws_torre.max_row  # La última fila de datos
        
        # Aplicar formato de color rojo claro a las filas con "CANCELADO AL INICIAR" en la segunda tabla
        for row_num in range(second_torre_table_first_row + 1, second_torre_table_last_row + 1):
            tipo_fin_cell = ws_torre[f"G{row_num}"]  # Columna G es "Tipo de Fin"
            if tipo_fin_cell.value == "CANCELADO AL INICIAR":
                # Aplicar color de fondo rojo claro a toda la fila
                for col_num in range(1, 14):  # Columnas A-M (13 columnas para la segunda tabla)
                    cell = ws_torre.cell(row=row_num, column=col_num)
                    cell.fill = light_red_fill
        
        # Agregar la segunda tabla de torre
        second_torre_table = Table(displayName="ResumenProductividadTorreCliente", ref=f"A{second_torre_table_first_row}:M{second_torre_table_last_row}")
        second_torre_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        second_torre_table.tableStyleInfo = second_torre_style
        ws_torre.add_table(second_torre_table)
        
        # Ajustar ancho de columnas para ambas hojas (excepto columna B en ws que ya se ajustó)
        for sheet in [ws, ws_torre]:
            for col in sheet.columns:
                max_data_length = 0
                max_header_length = 0
                column = None
                is_header_multiline = False
                
                for cell in col:
                    if hasattr(cell, "column_letter"):
                        column = cell.column_letter
                    else:
                        continue
                        
                    try:
                        if cell.value:
                            cell_str = str(cell.value)
                            
                            # Detectar si es un encabezado (filas 5-8 típicamente contienen encabezados)
                            if 5 <= cell.row <= 10 and ('[' in cell_str or 'Tiempo' in cell_str or 'Nivel' in cell_str or 'Peso' in cell_str or 'ID' in cell_str):
                                # Es un encabezado
                                if '\n' in cell_str:
                                    lines = cell_str.split('\n')
                                    max_header_length = max(len(line) for line in lines)
                                    is_header_multiline = True
                                else:
                                    max_header_length = len(cell_str)
                            else:
                                # Es dato normal
                                if cell.row > 10:  # Evitar títulos y fechas
                                    max_data_length = max(max_data_length, len(cell_str))
                    except:
                        pass
                
                if column:
                    # Establecer ancho fijo de 25 para la primera columna (A)
                    if column == "A":
                        final_width = 25
                    # Saltar columna B en hoja ws porque ya se ajustó específicamente
                    elif column == "B" and sheet == ws:
                        continue  # No modificar, ya se ajustó arriba
                    else:
                        # Calcular ancho basándose en el contenido más largo (header vs datos)
                        content_width = max(max_header_length, max_data_length)
                        
                        # Establecer rangos específicos por tipo de columna
                        if column == "B":
                            # Para otras hojas, usar la lógica normal
                            final_width = min(max(content_width + 2, 10), 25)
                        elif "Torres equivalentes" in str(sheet[column + "7"].value or "") or any("Torres equivalentes" in str(cell.value or "") for cell in sheet[column]):
                            final_width = 19  # Ancho específico para "Torres equivalentes"
                        elif "Kilos por hora" in str(sheet[column + "7"].value or "") or any("Kilos por hora" in str(cell.value or "") for cell in sheet[column]):
                            final_width = 19  # Ancho específico para "Kilos por hora"
                        elif "Tiempo" in str(max_header_length) or "MM:SS" in str(max_header_length):
                            final_width = 12  # Columnas de tiempo
                        elif is_header_multiline:
                            # Para encabezados multilínea, usar la línea más larga + pequeño margen
                            final_width = min(max(max_header_length + 2, 10), 20)
                        elif "[kg]" in str(sheet[column + "7"].value or "") or "Peso" in str(sheet[column + "7"].value or ""):
                            final_width = min(max(content_width + 2, 12), 16)  # Columnas de peso
                        elif "[seg]" in str(sheet[column + "7"].value or ""):
                            final_width = 21  # Segundos/Nivel
                        elif "%" in str(sheet[column + "7"].value or ""):
                            final_width = 12  # Porcentajes
                        else:
                            final_width = min(max(content_width + 2, 10), 25)  # Otras columnas
                    
                    if column != "B" or sheet != ws:  # Solo aplicar si no es la columna B en ws
                        sheet.column_dimensions[column].width = final_width

        wb.save(file_path)
        logger.info(f"Archivo Excel generado en: {file_path}")
        return True  # Retornar True indicando que se generó el Excel exitosamente
    except Exception as e:
        logger.error(f"Error exportando a Excel: {e}")
        raise